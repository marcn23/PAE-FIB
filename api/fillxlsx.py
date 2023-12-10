import os
import pandas as pd
import sys
from xlutils.copy import copy
import openpyxl


org = sys.argv[1]
csv_filename = str(org) + ".csv"
    
csvin = os.path.normpath(os.path.join(os.getcwd(),'in', 'CSVs', 'XLSX', csv_filename))
    
data = pd.read_csv(csvin)

xlsxin = os.path.normpath(os.path.join(os.getcwd(),'in',"SGAE_ResumAutoliquidacions.xlsx"))
xlsxout = os.path.normpath(os.path.join(os.getcwd(),'out','XLSX',"SGAE_ResumAutoliquidacions_" + org + ".xlsx"))

xfile = openpyxl.load_workbook(xlsxin)

sheet = xfile["Autoliquidacions"]

i=4; 
for j, rows in data.iterrows():
    # Use the worksheet object to write
    print(j)
    # data via the write() method.
    sheet.cell(row=i, column=1).value = rows['H. PROGRAMA']
    sheet.cell(row=i, column=2).value = rows['NOM ORQUESTRA O COBLA']
    sheet.cell(row=i, column=3).value = rows['DATA INICI']
    sheet.cell(row=i, column=4).value = rows['DATA FI']
    sheet.cell(row=i, column=5).value = rows['POBLACIÓ']
    sheet.cell(row=i, column=6).value = rows['PROVÍNCIA']
    sheet.cell(row=i, column=7).value = rows['LOCAL /CARRER/ESPAI']
    sheet.cell(row=i, column=8).value = rows['ESDEVENIMENT']
    sheet.cell(row=i, column=9).value = rows['NOM ENTITAT AUTOLIQUIDACIÓ']

    i+=1

xfile.save(xlsxout)